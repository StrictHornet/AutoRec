import tkinter as tk
from openpyxl import Workbook
from openpyxl import load_workbook

############# TASKS ###################
'''
Optimising application options
1. Make reconciliation sheet for each year its own thing
2. Use excel inbuilt search function to search fro each element 
3. Intuitively reduce number of iterations
4. Combine duplicate loop with rconcile loop
'''
# PUTTING EVERYTHING IN ONE CONSTANT FILE
# CREATE NEW FILE, COPY REC FILE, REC
## DESIGN ERROR TEXT FIELD
## DESIGN FUNCTION THAT PRINTS ERROR TO TEXT FIELD
# PRINT ERRORS TO TEXT FIELD
# "COMPLETED RECONCILIATION" TEXTFIELD
# FUNCTION THAT CONTROLS STATUS REPORT

#FUNCTION THAT RECONCILES SURVEY DOCUMENT#

def reconcile():
    file_directory = ent_directory.get()
    frm_status = tk.Frame(padx = 10,pady = 10)

    lbl_status = tk.Label(master = frm_status, 
                                            width=35, 
                                            bg="white", 
                                            padx=10, 
                                            text="STATUS REPORT")
    lbl_status.pack(fill=tk.X)

    txt_status = tk.Text(master=frm_status, width=55)
    txt_status.pack()

    frm_status.pack()

    errorCode = "Didn't work"
    app = (ord(approvalColumn.lower()) - 96) - 3
    com = (ord(commentColumn.lower()) - 96) - 3
    #The subtracted 3 is the distance between the beginning of the ir that
    #starts from the third column plus the one unit diffrence in the ir[] array
    workbook = load_workbook(filename=file_directory)

# For loop for search and cell assignment
    for ir in workbook["RecNew"].iter_rows(min_row=2, min_col=3):
        if ir[0].value is None:
                    continue
        for sheet in workbook:
            for row in sheet.iter_rows(min_row=2, min_col=8):
                try:
                    if ir[0].value == row[0].value:
                        if row[6].value != "NIL":
                            ir[app].value = "Failed" 
                            ir[com].value = row[6].value 
                            break  # Since IR has been found loop should break to next IR
                        else:
                            ir[app].value = "OK"
                            break
                except:
                    # errorCode = "Didn't work!"
                    # txt_status.insert(tk.END, f"\n{errorCode}")
                    break
            # else:
            #     continue
            break  # Break the outer loop

# For loop for duplicate search
    # for ir in workbook["RecNew"].iter_rows(min_row=2, min_col=3):
    #     for sheet in workbook:
    #         for row in sheet.iter_rows(min_row=2, min_col=3):
    #             try:
    #                 if ir[0].value == row[0].value:
    #                     if row[app].value == "Approved":
    #                         ir[app].value = "DUPLICATE" 
    #                         break  # Since IR has been found loop should break to next IR
    #             except:
    #                 print("Didn't work!")
    #         else:
    #             continue
    #         break  # Break the outer loop
    txt_status.insert(tk.END, "COMPLETED")
    workbook.save(filename=file_directory)

#TKINTER 
win = tk.Tk()
win.title("AutoRec")

frm_head = tk.Frame(bg="yellow")
frm_head.pack(fill=tk.X)

lbl_greeting = tk.Label(relief=tk.RIDGE, 
                                    height=2, 
                                    bg="#641822",
                                    text="Inq SLA Monthly Reconciler", 
                                    fg="white",
                                    master=frm_head)
lbl_greeting.pack(fill=tk.X)

frm_body = tk.Frame()
frm_body.pack(padx=15, pady=15)

approvalColumn = "X"
lbl_approvalColumn = tk.Label(text = "inq. Approval Column Alphabet Identifier", master = frm_body)
lbl_approvalColumn.grid(row=0, column=0)
ent_approvalColumn = tk.Entry(master = frm_body, width = 2)
ent_approvalColumn.insert(0, approvalColumn)
approvalColumn = ent_approvalColumn.get()
ent_approvalColumn.grid(row=0, column=1)

commentColumn = "Y"
lbl_commentColumn = tk.Label(text = "inq. Comment Column Alphabet Identifier", master = frm_body)
lbl_commentColumn.grid(row=1, column=0)
ent_commentColumn = tk.Entry(master = frm_body, width = 2)
ent_commentColumn.insert(0, commentColumn)
commentColumn = ent_commentColumn.get()
ent_commentColumn.grid(row=1, column=1)

frm_directory = tk.Frame(padx = 10,pady = 10)

lbl_directory = tk.Label(master = frm_directory, 
                                            width=35, 
                                            bg="gray", 
                                            padx=10, 
                                            text="DIRECTORY PATH")
lbl_directory.pack()

file_directory = "C:/Users/okosu/Google Drive/Work/VODACOM MAIN COPY OF SURVEY SPREADSHEET.xlsx"
ent_directory = tk.Entry(master=frm_directory, width=55)
ent_directory.insert(0, file_directory)
ent_directory.pack()

frm_directory.pack()

btn_reconcile = tk.Button(text="Reconcile", bg="#a00008", width=15, height=3, command=reconcile)
btn_reconcile.pack(pady=5)

win.mainloop()